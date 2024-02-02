from openpyxl import load_workbook


import os
import shutil

import re

from progress.bar import Bar

from datetime import datetime

#verifica si es un numero
def es_numero(num):
    try:
        float(num)
        return True
    except ValueError:
        return False



def leerArtic():
    #PASA EL ARCHIVO DE LOS ARTICULOS DE SISTEMA A UN ARCHIVO DE TEXTO FACIL DE LEER
    shutil.copy("../SIAAC3/ARTIC.DBF","DB/ARTIC.DBF")
    file = open("DB/ARTIC.DBF", errors="ignore")
    articdb = open("DB//articDB.txt", "w")
    for i in range(0,6):
    #se descarta las pimeras lineas
        linea = file.readline()


    linea = file.readline(2)
    linea = file.readline(200)
    while linea != "":

        result = re.findall("[A-Z]$", linea)
        if result:
            lineaAux = linea[199]
            linea = linea[:-132] + " " + linea[68:-1] + '\n'
            articdb.write(linea[:-189].lstrip('\x00').lstrip() + linea[89:])
            linea = lineaAux + file.readline(199)
        else:
            articdb.write(linea[:-188].lstrip('\x00').lstrip() + ' ' + linea[88:] + '\n')
            linea = file.readline(200)

    file.close()
    articdb.close()

def buscarPrecio(cod, lista_num):
    #BUSCA POR CODIGO EL PRECIO DEL ARTICULO 
    
    long_precio = 11
    if lista_num == 1:
        inicioPrecio = 22
    if lista_num == 5:
        inicioPrecio = 68

    finprecio = inicioPrecio + long_precio
    cod = cod.upper().strip()
    file = open("DB//articDB.txt")

    
    
    for linea in file:

        #result = re.findall(patron, linea)
        
        if cod == linea[:11].strip():
            precio = linea[inicioPrecio:finprecio]
           
            return precio.strip(" ")

    print(f"\n\n*********************************************\n ERROR: codigo {cod} no encontrado reviselo\n*********************************************\n")    
    return -1


def actualizarLista(bExcel, lista_num):
    #recorre una lista de precios, obtiene los codigo, los busca en articDB.txt y actualiza el precio
    try:
        sh = bExcel['Hoja1']
    except:
        print ("Error no exixte la Hoja1 en el archivo excel!!!!")
   
    cell=""
    i=0
    columna = 1

    now = datetime.now()    
    sh['A1'] = now

    maxFila = sh.max_row

    for i in range(1,maxFila):
        cell=str(sh.cell(row=i,column=columna).value).upper()
    
        celda = str(cell).upper()
        
        
        if celda == "COD" or celda == "COD.":
            i += 1
            cell=str(sh.cell(row=i,column=columna).value).upper()
            result = re.findall(f"[A-Z]-", cell)

            while result:
                precioActual = str(sh.cell(row=i,column=columna+3).value)
                if es_numero(precioActual):
                    actualizarPrecio(bExcel,i,cell,lista_num)
                    
                i+=1
                cell=sh.cell(row=i,column=columna).value
                if cell:                    
                    result = re.findall(f"[A-Z]-", cell.upper())
                else:
                    result = 0



def actualizarPrecio (wb,row,cod,lista_num):
    #ACTUALIZA EL PRECIO EN EL EXCEL
    
    sheet = wb['Hoja1']
    cell = 'D' + str(row)
    precio = buscarPrecio(cod, lista_num)
    if precio != -1:
        sheet[cell] = float(precio)
    

def getListas():
    #CREA UNA LISTA CON TODOS LOS ARCHVOS .XLSX EN EL MISMO DIRECTORIO DE EJECUCON
    archivos = os.listdir('./')
    lista = []
    
    for line in archivos:
        result = re.findall("\S.xlsx", line)
        if result:
            lista.append(line)
            
    return lista

def listar_archivos_directorio(directorio): # obtiene de cada lista la ruta completa del drive
    ruta_listas = {}
    for raiz, directorios, archivos in os.walk(directorio):
        for archivo in archivos:
            if archivo[-5:] == ".xlsx": # si es un archivo excel
                ruta_xlsx = os.path.join(raiz, archivo)
                end_index = ruta_xlsx[::-1].find("\\") # quitamos el nombre de la lista de la ruta
                end_index = len(ruta_xlsx) - end_index
                ruta_listas[archivo] = ruta_xlsx[:end_index]
    return ruta_listas

def copy_to_local(rute, ma_mi):# copia en la carpeta tradicional las listas
    if ma_mi == "ma":
        rute += "LISTA MAYORISTA/"
    elif ma_mi == "mi":
        rute += "LISTA MINORISTA/"
    else:
        raise ValueError("Solo acepta ma o mi")
    listas_to_copy = getListas() # OBTIENE LAS LISTA A COPIAR
    if listas_to_copy != []:
        for lista in listas_to_copy:
            try:
                shutil.copyfile(lista, rute + lista)
            except Exception as e:
                print(f"\nNo se pudo copiar la lista {lista}")
                print(f"Error[{type(e).__name__}]: {e}")


def copy_to_drive(ma_mi, bar): # copia y reemplaza las listas en la carpeta q esta sincronizada con el drive
    ruta_drive = R"C:\RED\LISTAS Drive\\" + ma_mi.upper()
    listas_to_copy = getListas() # OBTIENE LAS LISTA A COPIAR
    drive = listar_archivos_directorio(ruta_drive) # RUTAS ABSOLUTA DE LAS LISTAS DEL DRIVE
    if listas_to_copy != [] and drive != {}:
        for lista in listas_to_copy:
            try:
                shutil.copy2(lista, drive[lista])
            except KeyError:
                print(f"El archivo {lista} no existe en el drive") # si no existe en drive no hace lo copia
                print(f"Si desea agregarlo al drive copie el archivo en {ruta_drive}")
            except Exception as e:
                print(F"Error: [{type(e).__name__}] no se puedo copiar {lista} en {drive[lista]}")
            bar.next()

        



'''     while lista_ma_mi == "":
        os.system("cls")   
        lista_ma_mi = input(f"{msj}Ingrese 'MA' para mayorista o 'MI' para minorista: ")

        if lista_ma_mi.upper() == 'MI':
            lista_num = 1
        elif lista_ma_mi.upper() == 'MA':
            lista_num = 5
        else:
            lista_ma_mi = ""
            msj = "Opcion no valida\n"
'''
if __name__ == '__main__':

    leerArtic()
    rute_local = "Y:/Lista de Precio/"
    listas =getListas()

    num_listas = len(listas)

    bar1 = Bar("Actualizando listas minoristas:", max=num_listas)
    lista_ma_mi = ""
    lista_num = 1
    msj = ""

    
    

    #hay
    
    for arch in listas:
        try:
            book = load_workbook(filename= arch)
        except:
            print(F"\n\nERROR NO SE PUEDE ABRIR LA LISTA{arch}\nINTENTE CERRAR LAS LISTA E INTETELO NUEVAMENTE")
            
        
        actualizarLista(book, lista_num)

        
        try:
            book.save(f"./{arch}")
        except:
            print(f"\n\nError no se puedo guardar el archivo {arch}. intetnte cerrar todos los archivos excel e intentelo nuevamente")
        
        bar1.next()
    bar1.finish()
    bar2 = Bar("Copiando listas:", max=num_listas)

    copy_to_drive("mi", bar2)
    copy_to_local(rute_local,"mi")
    bar2.finish()

    bar3 = Bar("Actualizando listas mayorista:", max=num_listas)
    lista_num = 5
    for arch in listas:
        try:
            book = load_workbook(filename= arch)
        except:
            print(F"\n\nERROR NO SE PUEDE ABRIR LA LISTA{arch}\nINTENTE CERRAR LAS LISTA E INTETELO NUEVAMENTE")
            
        
        actualizarLista(book, lista_num)

        
        try:
            book.save(f"./{arch}")
        except:
            print(f"\n\nError no se puedo guardar el archivo {arch}. intetnte cerrar todos los archivos excel e intentelo nuevamente")
        
        bar3.next()
    bar3.finish()
    bar4 = Bar("Copiando listas:", max=num_listas)

    copy_to_drive("ma", bar4)
    copy_to_local(rute_local,"ma")
    bar2.finish()

    input("\n\n--PRECIONE ENTER PARA SALIR--")

    

    
